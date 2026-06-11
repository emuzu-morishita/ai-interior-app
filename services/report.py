"""生成結果（提案アイテム・おすすめ商品・完成予想図）を、画像を base64 で
埋め込んだ単一HTMLファイルにまとめるモジュール。

ブラウザのリロードで st.session_state が消えても手元に残せるよう、
外部リソースに依存しない自己完結したHTMLを生成する。
"""
import base64
import io
import html as html_lib

from services.i18n import t


def _image_data_uri(image) -> str:
    """PIL Image を PNG の data URI 文字列に変換する。"""
    buf = io.BytesIO()
    image.save(buf, format="PNG")
    b64 = base64.b64encode(buf.getvalue()).decode("ascii")
    return f"data:image/png;base64,{b64}"


def _product_cards(products, lang: str) -> str:
    """おすすめ商品を固定サイズのHTMLカード列としてレンダリングする。"""
    cards = []
    for p in products:
        name = html_lib.escape(p.name[:60] + ("…" if len(p.name) > 60 else ""))
        url = html_lib.escape(p.url or "", quote=True)
        img = html_lib.escape(p.image_url or "", quote=True)
        src = html_lib.escape(p.source)
        img_div = (
            f'<span class="p-img" style="background-image:url(\'{img}\')"></span>'
            if img else '<span class="p-img"></span>'
        )
        review = ""
        if p.review_average > 0:
            review = f'<span class="p-review">{t(lang, "review_fmt", avg=p.review_average, count=p.review_count)}</span>'
        cards.append(
            f'<a class="p-card" href="{url}" target="_blank" rel="noopener">'
            f'{img_div}'
            f'<span class="p-badge">🏷️ {src}</span>'
            f'<span class="p-name">{name}</span>'
            f'<span class="p-price">¥{p.price:,}</span>'
            f'{review}'
            f'</a>'
        )
    return f'<div class="p-row">{"".join(cards)}</div>'


def _budget_section_html(items, total: int, lang: str) -> str:
    """予算内訳を、自己完結したCSSバーのHTMLセクションとして返す。"""
    if not items or total <= 0:
        return ""
    rows = []
    for it in items:
        pct = round(it["price"] / total * 100)
        name = html_lib.escape(it["item_name"])
        rows.append(
            f'<div class="bd-row">'
            f'<span class="bd-name">{name}</span>'
            f'<span class="bd-track"><span class="bd-fill" style="width:{pct}%"></span></span>'
            f'<span class="bd-val">¥{it["price"]:,}<small>{pct}%</small></span>'
            f'</div>'
        )
    return (
        f'<section class="card"><h2>📊 {html_lib.escape(t(lang, "budget_breakdown"))}</h2>'
        f'<div class="bd">{"".join(rows)}</div></section>'
    )


def build_html_report(
    items,
    shopping_results,
    room_image,
    *,
    lang: str,
    room_label: str,
    taste: str,
    budget: int,
    generated_at: str,
    owned_items: str = "",
) -> str:
    """生成結果一式を単一HTMLファイル（画像埋め込み）にまとめて返す。"""
    total = sum(it["price"] for it in items)

    # --- 完成予想図 ---
    image_block = ""
    if room_image is not None:
        image_block = (
            f'<section class="card preview">'
            f'<h2>🖼️ {html_lib.escape(t(lang, "preview_header"))}</h2>'
            f'<img src="{_image_data_uri(room_image)}" alt="room preview">'
            f'<p class="cap">{html_lib.escape(t(lang, "preview_caption"))}</p>'
            f'</section>'
        )

    # --- 提案条件 ---
    meta_rows = "".join(
        f'<div class="meta-row"><span class="k">{html_lib.escape(k)}</span>'
        f'<span class="v">{html_lib.escape(v)}</span></div>'
        for k, v in _conditions(lang, room_label, taste, budget, total, generated_at, owned_items)
    )

    budget_block = _budget_section_html(items, total, lang)

    # --- アイテム一覧 ---
    item_blocks = []
    for i, it in enumerate(items, 1):
        name = html_lib.escape(it["item_name"])
        reason = html_lib.escape(it["reason"])
        place_block = ""
        if it.get("placement"):
            place_block = (
                f'<p class="item-place">📐 <b>{html_lib.escape(t(lang, "placement_label"))}</b> '
                f'{html_lib.escape(it["placement"])}</p>'
            )
        products = shopping_results.get(it["item_name"], [])
        rec_block = ""
        if products:
            rec_block = (
                f'<div class="rec-h">{html_lib.escape(t(lang, "recommend_header"))}</div>'
                f'{_product_cards(products, lang)}'
            )
        item_blocks.append(
            f'<div class="item">'
            f'<div class="item-head">'
            f'<span class="item-no">{i}</span>'
            f'<span class="item-name">{name}</span>'
            f'<span class="item-price">¥{it["price"]:,}</span>'
            f'</div>'
            f'<p class="item-reason">{reason}</p>'
            f'{place_block}'
            f'{rec_block}'
            f'</div>'
        )

    return f"""<!DOCTYPE html>
<html lang="{html_lib.escape(lang, quote=True)}">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>AI Interior Coordinator</title>
<style>
* {{ box-sizing: border-box; }}
body {{ margin: 0; padding: 24px; background: #fbfaf7; color: #2b2b28;
  font-family: -apple-system, "Segoe UI", "Hiragino Kaku Gothic ProN", "Yu Gothic", "Microsoft YaHei", "Malgun Gothic", Meiryo, sans-serif; }}
.wrap {{ max-width: 880px; margin: 0 auto; }}
.hero {{ background: linear-gradient(120deg, #4d7c0f 0%, #84a98c 100%);
  border-radius: 22px; padding: 28px 34px; color: #fff; box-shadow: 0 10px 28px rgba(77,124,15,.2); }}
.hero h1 {{ margin: 0; font-size: 30px; font-weight: 800; }}
.hero p {{ margin: 8px 0 0; opacity: .93; font-size: 14px; }}
.card {{ background: #fff; border: 1px solid #e7e3d8; border-radius: 16px;
  padding: 20px 24px; margin-top: 18px; box-shadow: 0 2px 12px rgba(0,0,0,.04); }}
h2 {{ font-size: 18px; color: #4d7c0f; margin: 0 0 14px; }}
.meta-row {{ display: flex; justify-content: space-between; padding: 9px 0;
  border-bottom: 1px dashed #e7e3d8; font-size: 14px; }}
.meta-row:last-child {{ border-bottom: none; }}
.meta-row .k {{ color: #7a7468; }}
.meta-row .v {{ font-weight: 700; }}
.preview img {{ width: 100%; border-radius: 12px; display: block; }}
.preview .cap {{ font-size: 12px; color: #9a8f7d; margin: 8px 0 0; text-align: center; }}
.item {{ border-top: 1px solid #efece3; padding: 18px 0; }}
.item:first-of-type {{ border-top: none; }}
.item-head {{ display: flex; align-items: center; gap: 10px; }}
.item-no {{ flex: none; width: 26px; height: 26px; border-radius: 50%; background: #4d7c0f;
  color: #fff; font-size: 13px; font-weight: 700; display: flex; align-items: center; justify-content: center; }}
.item-name {{ font-weight: 700; font-size: 16px; flex: 1; }}
.item-price {{ font-weight: 800; color: #4d7c0f; white-space: nowrap; }}
.item-reason {{ margin: 10px 0 0; font-size: 14px; line-height: 1.7; color: #4a4a44; }}
.item-place {{ margin: 8px 0 0; font-size: 13px; line-height: 1.6; color: #4a4a44;
  background: #f7f5ed; border-left: 3px solid #4d7c0f; border-radius: 0 8px 8px 0; padding: 8px 12px; }}
.rec-h {{ font-size: 12px; font-weight: 700; color: #4d7c0f; margin: 14px 0 9px; }}
.p-row {{ display: flex; flex-wrap: wrap; gap: 12px; }}
.p-card {{ width: 150px; border: 1px solid #e7e3d8; border-radius: 12px; padding: 9px;
  text-decoration: none; color: inherit; background: #fff; }}
.p-img {{ display: block; width: 100%; aspect-ratio: 1/1; border-radius: 8px; background: #f1efe8;
  background-size: cover; background-position: center; }}
.p-badge {{ display: inline-block; font-size: 10px; font-weight: 700; color: #4d7c0f;
  background: #eef3e3; border-radius: 6px; padding: 2px 7px; margin-top: 8px; }}
.p-name {{ display: block; font-size: 11px; line-height: 1.35; height: 30px; overflow: hidden; margin-top: 5px; color: #3a3a36; }}
.p-price {{ display: block; font-size: 15px; font-weight: 800; margin-top: 5px; }}
.p-review {{ display: block; font-size: 10px; color: #9a8f7d; margin-top: 2px; }}
.bd {{ margin-top: 2px; }}
.bd-row {{ display: flex; align-items: center; gap: 10px; margin: 9px 0; font-size: 13px; }}
.bd-name {{ width: 34%; color: #4a4a44; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }}
.bd-track {{ flex: 1; height: 10px; background: #eef0e8; border-radius: 999px; overflow: hidden; }}
.bd-fill {{ display: block; height: 100%; background: linear-gradient(90deg,#84a98c,#4d7c0f); border-radius: 999px; }}
.bd-val {{ width: 24%; text-align: right; font-weight: 700; color: #2b2b28; white-space: nowrap; }}
.bd-val small {{ color: #9a8f7d; font-weight: 600; margin-left: 5px; }}
footer {{ text-align: center; color: #9a8f7d; font-size: 11px; margin: 26px 0 6px; }}
</style>
</head>
<body>
  <div class="wrap">
    <header class="hero">
      <h1>🛋️ AI Interior Coordinator</h1>
      <p>{html_lib.escape(t(lang, "caption"))}</p>
    </header>
    <section class="card">
      <h2>📋 {html_lib.escape(t(lang, "report_conditions"))}</h2>
      {meta_rows}
    </section>
    {image_block}
    <section class="card">
      <h2>🛋️ {html_lib.escape(t(lang, "items_header"))}</h2>
      {"".join(item_blocks)}
    </section>
    {budget_block}
    <footer>AI Interior Coordinator · {html_lib.escape(generated_at)}</footer>
  </div>
</body>
</html>"""


def _conditions(lang, room_label, taste, budget, total, generated_at, owned_items=""):
    """各形式で共通して使う「提案条件」の (ラベル, 値) リストを返す。"""
    rows = [
        (t(lang, "room_size_label"), room_label),
        (t(lang, "taste_label"), taste),
        (t(lang, "budget_metric"), f"¥{budget:,}"),
        (t(lang, "m_total"), f"¥{total:,}"),
        (t(lang, "report_generated"), generated_at),
    ]
    owned = (owned_items or "").strip()
    if owned:
        rows.append((t(lang, "owned_label"), owned))
    return rows


# 言語コード → reportlab 同梱のCIDフォント名（追加フォントファイル不要）
_PDF_CID_FONTS = {
    "ja": "HeiseiKakuGo-W5",
    "ko": "HYSMyeongJo-Medium",
    "zh": "STSong-Light",
}


def _register_pdf_font(lang: str) -> str:
    """言語に応じたCIDフォントを登録し、フォント名を返す。失敗時は Helvetica。"""
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont

    name = _PDF_CID_FONTS.get(lang)
    if not name:
        return "Helvetica"
    try:
        if name not in pdfmetrics.getRegisteredFontNames():
            pdfmetrics.registerFont(UnicodeCIDFont(name))
        return name
    except Exception:
        return "Helvetica"


def build_pdf_report(
    items,
    shopping_results,
    room_image,
    *,
    lang: str,
    room_label: str,
    taste: str,
    budget: int,
    generated_at: str,
    owned_items: str = "",
) -> bytes:
    """生成結果を、見た目を保ったままのPDFバイト列にして返す。"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image as RLImage,
    )

    font = _register_pdf_font(lang)
    total = sum(it["price"] for it in items)

    green = colors.HexColor("#4d7c0f")
    dark = colors.HexColor("#2b2b28")
    grey = colors.HexColor("#7a7468")
    body_col = colors.HexColor("#4a4a44")
    line_col = colors.HexColor("#e7e3d8")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4, leftMargin=18 * mm, rightMargin=18 * mm,
        topMargin=16 * mm, bottomMargin=16 * mm, title="AI Interior Coordinator",
    )

    h1 = ParagraphStyle("h1", fontName=font, fontSize=18, textColor=colors.white, leading=22)
    sub = ParagraphStyle("sub", fontName=font, fontSize=9, textColor=colors.white, leading=13)
    h2 = ParagraphStyle("h2", fontName=font, fontSize=13, textColor=green, leading=18,
                        spaceBefore=12, spaceAfter=6)
    key = ParagraphStyle("key", fontName=font, fontSize=10, textColor=grey, leading=15)
    val = ParagraphStyle("val", fontName=font, fontSize=10, textColor=dark, leading=15)
    ih = ParagraphStyle("ih", fontName=font, fontSize=12, textColor=dark, leading=16, spaceBefore=10)
    body = ParagraphStyle("body", fontName=font, fontSize=10, textColor=body_col, leading=15, spaceBefore=3)
    place = ParagraphStyle("place", fontName=font, fontSize=9, textColor=green, leading=14, spaceBefore=3)
    rec = ParagraphStyle("rec", fontName=font, fontSize=9, textColor=green, leading=13, spaceBefore=6)
    prod = ParagraphStyle("prod", fontName=font, fontSize=9, textColor=dark, leading=14, leftIndent=10)
    cap = ParagraphStyle("cap", fontName=font, fontSize=8, textColor=grey, leading=11, alignment=1)

    def esc(s):
        return html_lib.escape(str(s))

    story = []

    # --- ヘッダーバナー ---
    banner = Table(
        [[Paragraph("AI Interior Coordinator", h1)],
         [Paragraph(esc(t(lang, "caption")), sub)]],
        colWidths=[doc.width],
    )
    banner.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), green),
        ("LEFTPADDING", (0, 0), (-1, -1), 14),
        ("RIGHTPADDING", (0, 0), (-1, -1), 14),
        ("TOPPADDING", (0, 0), (0, 0), 12),
        ("BOTTOMPADDING", (0, -1), (-1, -1), 12),
        ("TOPPADDING", (0, 1), (-1, 1), 4),
    ]))
    story.append(banner)
    story.append(Spacer(1, 14))

    # --- 提案条件 ---
    story.append(Paragraph(esc(t(lang, "report_conditions")), h2))
    cond_rows = [[Paragraph(esc(k), key), Paragraph(esc(v), val)]
                 for k, v in _conditions(lang, room_label, taste, budget, total, generated_at, owned_items)]
    cond_tbl = Table(cond_rows, colWidths=[doc.width * 0.32, doc.width * 0.68])
    cond_tbl.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LINEBELOW", (0, 0), (-1, -1), 0.5, line_col),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(cond_tbl)

    # --- 完成予想図 ---
    if room_image is not None:
        img_buf = io.BytesIO()
        room_image.save(img_buf, format="PNG")
        img_buf.seek(0)
        iw, ih_px = room_image.size
        w = doc.width
        h = w * ih_px / iw
        max_h = 120 * mm
        if h > max_h:
            h = max_h
            w = h * iw / ih_px
        story.append(Paragraph(esc(t(lang, "preview_header")), h2))
        story.append(RLImage(img_buf, width=w, height=h))
        story.append(Paragraph(esc(t(lang, "preview_caption")), cap))

    # --- アイテム一覧 ---
    story.append(Paragraph(esc(t(lang, "items_header")), h2))
    for i, it in enumerate(items, 1):
        story.append(Paragraph(f"{i}. {esc(it['item_name'])} — ¥{it['price']:,}", ih))
        story.append(Paragraph(esc(it["reason"]), body))
        if it.get("placement"):
            story.append(Paragraph(f"{esc(t(lang, 'placement_label'))} {esc(it['placement'])}", place))
        products = shopping_results.get(it["item_name"], [])
        if products:
            story.append(Paragraph(esc(t(lang, "recommend_header")), rec))
            for p in products:
                pname = p.name[:50] + ("…" if len(p.name) > 50 else "")
                url = html_lib.escape(p.url or "", quote=True)
                link = (
                    f'• <a href="{url}" color="#4d7c0f">{esc(pname)}</a> '
                    f'— ¥{p.price:,}（{esc(p.source)}）'
                )
                story.append(Paragraph(link, prod))
        story.append(Spacer(1, 6))

    # --- 予算内訳（横棒グラフ） ---
    if total > 0:
        from reportlab.graphics.shapes import Drawing, Rect, String

        story.append(Paragraph(esc(t(lang, "budget_breakdown")), h2))
        row_h = 18
        label_w = doc.width * 0.30
        bar_max = doc.width * 0.46
        d = Drawing(doc.width, row_h * len(items))
        for k, it in enumerate(items):
            y = (len(items) - 1 - k) * row_h + 4
            frac = it["price"] / total
            label = it["item_name"]
            if len(label) > 14:
                label = label[:13] + "…"
            d.add(String(0, y, label, fontName=font, fontSize=8, fillColor=dark))
            d.add(Rect(label_w, y - 1, bar_max, 9, fillColor=line_col, strokeColor=None))
            d.add(Rect(label_w, y - 1, max(1.0, bar_max * frac), 9, fillColor=green, strokeColor=None))
            d.add(String(label_w + bar_max + 6, y, f"¥{it['price']:,} ({round(frac * 100)}%)",
                         fontName=font, fontSize=8, fillColor=dark))
        story.append(d)

    doc.build(story)
    return buf.getvalue()


def build_excel_report(
    items,
    shopping_results,
    room_image,
    *,
    lang: str,
    room_label: str,
    taste: str,
    budget: int,
    generated_at: str,
    owned_items: str = "",
) -> bytes:
    """生成結果を、条件・アイテム表・おすすめ商品・完成予想図を含むExcelにして返す。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.drawing.image import Image as XLImage

    total = sum(it["price"] for it in items)
    GREEN = "4D7C0F"
    HEAD_BG = "EEF3E3"
    white = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="E7E3D8")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap_top = Alignment(wrap_text=True, vertical="top")

    wb = Workbook()
    ws = wb.active
    ws.title = (t(lang, "items_header") or "Proposal")[:28]

    # タイトル帯
    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = "AI Interior Coordinator"
    c.font = Font(size=16, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=GREEN)
    c.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 30

    # 提案条件
    r = 3
    for k, v in _conditions(lang, room_label, taste, budget, total, generated_at, owned_items):
        ws.cell(r, 1, k).font = Font(bold=True, color="7A7468")
        ws.cell(r, 2, v)
        r += 1
    r += 1

    # アイテム表
    headers = ["#", t(lang, "chart_item"), t(lang, "chart_price"),
               t(lang, "reason_label"), t(lang, "placement_label")]
    item_header_row = r
    for ci, htext in enumerate(headers, 1):
        cell = ws.cell(r, ci, htext)
        cell.font = Font(bold=True, color=GREEN)
        cell.fill = PatternFill("solid", fgColor=HEAD_BG)
        cell.border = border
    r += 1
    first_item_row = r
    for i, it in enumerate(items, 1):
        ws.cell(r, 1, i).border = border
        ws.cell(r, 2, it["item_name"]).border = border
        pc = ws.cell(r, 3, it["price"])
        pc.number_format = "¥#,##0"
        pc.border = border
        rc = ws.cell(r, 4, it["reason"])
        rc.alignment = wrap_top
        rc.border = border
        plc = ws.cell(r, 5, it.get("placement", ""))
        plc.alignment = wrap_top
        plc.border = border
        r += 1
    last_item_row = r - 1
    # 合計行
    ws.cell(r, 2, t(lang, "m_total")).font = Font(bold=True)
    tc = ws.cell(r, 3, total)
    tc.number_format = "¥#,##0"
    tc.font = Font(bold=True, color=GREEN)
    r += 2

    # 予算内訳チャート（横棒）：価格列を参照したネイティブグラフを表の右に配置
    if items and total > 0:
        from openpyxl.chart import BarChart, Reference

        chart = BarChart()
        chart.type = "bar"
        chart.title = t(lang, "budget_breakdown")
        chart.legend = None
        chart.height = 1.6 * len(items) + 2
        chart.width = 14
        data = Reference(ws, min_col=3, min_row=item_header_row, max_row=last_item_row)
        cats = Reference(ws, min_col=2, min_row=first_item_row, max_row=last_item_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, f"G{item_header_row}")

    # おすすめ商品
    has_products = any(shopping_results.get(it["item_name"]) for it in items)
    if has_products:
        ws.cell(r, 1, t(lang, "recommend_header")).font = Font(size=12, bold=True, color=GREEN)
        r += 1
        rec_headers = [t(lang, "chart_item"), t(lang, "col_product"),
                       t(lang, "chart_price"), t(lang, "col_mall"), ""]
        for ci, htext in enumerate(rec_headers, 1):
            cell = ws.cell(r, ci, htext)
            cell.font = Font(bold=True, color=GREEN)
            cell.fill = PatternFill("solid", fgColor=HEAD_BG)
            cell.border = border
        r += 1
        for it in items:
            for p in shopping_results.get(it["item_name"], []):
                ws.cell(r, 1, it["item_name"]).border = border
                ws.cell(r, 2, p.name).border = border
                pc = ws.cell(r, 3, p.price)
                pc.number_format = "¥#,##0"
                pc.border = border
                ws.cell(r, 4, p.source).border = border
                link = ws.cell(r, 5, t(lang, "product_button"))
                if p.url:
                    link.hyperlink = p.url
                    link.font = Font(color="0563C1", underline="single")
                link.border = border
                r += 1

    # 列幅（E はアイテム表の配置アドバイスと商品表のリンク列を兼ねる）
    for col, width in {"A": 16, "B": 34, "C": 14, "D": 40, "E": 34}.items():
        ws.column_dimensions[col].width = width

    # 完成予想図（別シートに画像を埋め込み）
    if room_image is not None:
        ws_img = wb.create_sheet((t(lang, "preview_header") or "Preview")[:28])
        img_buf = io.BytesIO()
        room_image.save(img_buf, format="PNG")
        img_buf.seek(0)
        pic = XLImage(img_buf)
        # 横幅 約720px に収める
        scale = 720 / pic.width if pic.width else 1
        pic.width = int(pic.width * scale)
        pic.height = int(pic.height * scale)
        ws_img.add_image(pic, "A1")

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
